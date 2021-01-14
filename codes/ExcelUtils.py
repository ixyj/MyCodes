# coding=utf-8

# pip3 install openpyxl
import os, sys
import openpyxl

class ExcelUtils:
    @staticmethod
    def ReadSheetData(excelFile, sheetName = None, sheetOffset = 0):
        workBook = openpyxl.load_workbook(excelFile, data_only=True)
        sheet = workBook[sheetName] if sheetName else workBook.worksheets[sheetOffset]
        # return None for merged cells (except the first)
        for i in range(1, sheet.max_row+1):
            yield [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column+1)]

    @staticmethod
    def ReadAllSheetData(excelFile):
        workBook = openpyxl.load_workbook(excelFile, data_only=True)
        for sheet in workBook.worksheets:
            # return None for merged cells (except the first)
            yield sheet.title, [[r.value for r in row] for row in sheet.rows]

    @staticmethod
    def Write(file, data):  #data = {'sheet name': [data list]}
        wb = openpyxl.Workbook()
        index = 1
        for key,value in data.items():            
            sheet = wb.create_sheet(key, index=index)
            index += 1
            for row in value: sheet.append(row)
        wb.remove(wb.worksheets[0])
        wb.save(file)

# del 按sheet名称删除
# del wb['Data']

# 直接赋值可以改工作表的名称
# sheet = wb.active
# sheet.title = 'Sheet1'

# 直接给单元格赋值
# sheet['A1'] = '编号'

# 添加一行
# row = [1, 2, 3, 4, 5]
# sheet.append(row)

# sheet.delete_cols
# sheet.delete_rows