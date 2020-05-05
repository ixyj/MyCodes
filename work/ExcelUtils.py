# coding=utf-8

# pip3 install openpyxl

import openpyxl

class ExcelUtils:
    @staticmethod
    def ReadSheetData(excelFile, sheetName = None):
        workBook = openpyxl.load_workbook(excelFile, data_only=True)
        sheet = workBook[sheetName] if sheetName else workBook.worksheets[0]
        # return None for merged cells (except the first)
        for i in range(1, sheet.max_row+1):
            yield [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column+1)]   